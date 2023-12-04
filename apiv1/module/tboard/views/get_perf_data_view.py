import collections
import math
import os
import pathlib
import sys

import numpy as np
import openpyxl
import pandas as pd
import re
from decimal import Decimal, ROUND_HALF_UP

from django.http import FileResponse
from django.utils.encoding import escape_uri_path
from django.db.models.functions import Cast, TruncSecond
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.views import APIView
from rest_framework import serializers
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count, Avg, Max, CharField

from apiv1.core.model import CustomPatternCharField
from apiv1.core.response import reef_400_response, ReefResponse
from apiv1.core.utils import date_format_transverter
from apiv1.core.view.generic import AutoExecuteSerializerGenericAPIView
from apiv1.module.device.models import PhoneModel, RomVersion, Device
from apiv1.module.job.models import Job, JobTestArea
from apiv1.module.rds.models import Rds
from apiv1.module.tboard.models import TBoard
from reef import settings
from reef.settings import MEDIA_ROOT, RESOURCE_EXCEL_FILE_EXPORT_PATH, RESOURCE_EXCEL_FILE_EXPORT


class CheckPerfTboardDetailDataSerilizer(serializers.Serializer):
    tboard = serializers.SlugRelatedField(
        queryset=TBoard.objects.all(),
        slug_field='id'
    )

    devices = CustomPatternCharField(
        queryset=Device.objects.all(),
        required=False
    )


class PerfTboardDetailDataSerilizer(serializers.ModelSerializer):
    device_name = serializers.SerializerMethodField()
    device_id = serializers.SerializerMethodField()
    job_list = serializers.SerializerMethodField()
    author_name = serializers.SerializerMethodField()

    def get_device_name(self, obj):
        return obj.device.values_list('device_name', flat=True).order_by('id')

    def get_job_list(self, obj):
        return obj.job.all()

    def get_device_id(self, obj):
        return obj.device.values_list('id', flat=True).order_by('id')

    def get_author_name(self, obj):
        return obj.author.username

    class Meta:
        model = TBoard
        fields = ('id', 'author', 'board_name', 'repeat_time', 'end_time', 'device_name', 'job_list',
                  'device_id', 'author_name')


class CheckoutPhoneModelAndRomVersonSerilizer(serializers.Serializer):
    phone_model = serializers.SlugRelatedField(
        queryset=PhoneModel.objects.all(),
        many=True,
        slug_field='id'
    )

    rom_version = serializers.SlugRelatedField(
        queryset=RomVersion.objects.all(),
        slug_field='id'
    )


class PerfViewSerilizer(serializers.Serializer):
    tboard = serializers.IntegerField()
    job = serializers.IntegerField()
    devices = CustomPatternCharField(
        queryset=Device.objects.all(),
        required=False
    )


class XLSTboradSerilizer(serializers.Serializer):
    tboard = serializers.PrimaryKeyRelatedField(
        queryset=TBoard.objects.all()
    )

    job_res_rule = serializers.JSONField()


class PerfDataPreviewSerializer(serializers.Serializer):
    tboard = serializers.PrimaryKeyRelatedField(
        queryset=TBoard.objects.all()
    )


class CheckoutTboardSerilizer(serializers.Serializer):
    tboard = serializers.CharField(
        required=False
    )

    job = serializers.CharField(
        required=False
    )

    devices = CustomPatternCharField(
        queryset=Device.objects.all(),
        required=False
    )

    phone_model_obj = serializers.SlugRelatedField(
        queryset=PhoneModel.objects.all(),
        slug_field='id',
        required=False
    )

    rom_version_obj = serializers.SlugRelatedField(
        queryset=RomVersion.objects.all(),
        slug_field='id',
        required=False
    )

    model_rom_version = serializers.CharField(
        required=False,
    )

    sizer = serializers.ChoiceField(
        required=False,
        choices=('line_chart', 'table_chart', 'box_chart'),
        default='line_chart'
    )

    def validate_tboard(self, val: str):
        if not val:
            return None
        pattern = "^([0-9]+,)*[0-9]+$"
        if isinstance(val, str) and re.match(pattern, val):
            return [tboard_id for tboard_id in val.split(',')]
        raise ValidationError("devices invalid!")

    def validate_job(self, val: str):
        if not val:
            return None
        pattern = "^([0-9]+,)*[0-9]+$"
        if isinstance(val, str) and re.match(pattern, val):
            return [job_id for job_id in val.split(',')]
        raise ValidationError("devices invalid!")

    def validate_model_rom_version(self, val: str):
        if not val:
            return None
        if isinstance(val, str):
            return [[data for data in model_rom_data.split('.')] for model_rom_data in val.split(',')]
        else:
            return None


class JobSerilizer(serializers.ModelSerializer):
    test_area = serializers.SerializerMethodField()
    custom_tag = serializers.SerializerMethodField()

    def get_test_area(self, obj):
        results = []
        job_test_area_queryset = obj.test_area.all()
        for job_test_area in job_test_area_queryset:
            results.append({'id': job_test_area.id, 'description': job_test_area.description})
        return results

    def get_custom_tag(self, obj):
        results = []
        custom_tag_queryset = obj.custom_tag.all()
        for custom_tag in custom_tag_queryset:
            results.append({'id': custom_tag.id, 'custom_tag_name': custom_tag.custom_tag_name})
        return results

    class Meta:
        model = Job
        fields = ('job_name', 'test_area', 'custom_tag', 'updated_time', 'id')


class GetPerfTboardDetailData(APIView):

    def get(self, request):
        """
        Perf Data detail data
        Author: Goufuqiang
        Date: 2020/8/27
        :param request: &tboard=1
        :return:
        """
        serializer = CheckPerfTboardDetailDataSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        tboard = serializer.validated_data.get('tboard', None)
        device_queryset = serializer.validated_data.get('devices', [])
        serializer = PerfTboardDetailDataSerilizer(tboard)
        result = dict(serializer.data)
        result['job_data'] = []
        job_id_list = result.pop('job_list', [])
        for job in job_id_list:
            rds_queryset = Rds.objects.filter(tboard=tboard, job_id=job,
                                              job_assessment_value='0', device__in=device_queryset)
            # avg
            avg = rds_queryset.aggregate(job_duration_avg=Avg('job_duration')).get('job_duration_avg', 0)
            avg = "暂无数据" if avg is None else round(avg, 3)
            # max
            max = rds_queryset.aggregate(job_duration_max=Max('job_duration')).get('job_duration_max', 0)
            max = "暂无数据" if max is None else round(max, 3)
            # 中位数
            job_duration_list = rds_queryset.exclude(job_duration=None).values_list('job_duration', flat=True)
            median = round(np.median(job_duration_list), 3) if job_duration_list else "暂无数据"
            # success number
            success_num = rds_queryset.filter(job_assessment_value=0).count()
            # failed number
            failed_num = Rds.objects.filter(tboard=tboard, job_id=job, job_assessment_value='1').count()
            job_info = {'avg': avg, 'max': max, 'median': median, 'success_num': success_num, 'failed_num': failed_num,
                        'job_id': job.id, 'job_num': job.job_name}
            result['job_data'].append(job_info)
        return Response(result, status=status.HTTP_200_OK)


class GetPerfDataJobData(APIView):

    def get(self, request):
        """
        Perf Data get job page
        Author: Goufuqiang
        Date: 2020/8/27
        :param request: &tboard=1,2,3
        :return:
        """
        serializer = CheckoutTboardSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        tboard_id_list = serializer.validated_data.get('tboard', [])
        job_name__icontains = {
            'job_name__icontains': request.query_params['job_name__icontains']} if request.query_params.get(
            'job_name__icontains', None) else {}
        job_queryset = TBoard.objects.filter(id__in=tboard_id_list).distinct('job').values_list('job', flat=True)
        job_queryset = Job.objects.filter(id__in=list(job_queryset), **job_name__icontains)
        count = job_queryset.count()
        pagination = LimitOffsetPagination()
        job_queryset = pagination.paginate_queryset(job_queryset, request, view=None)
        serializer = JobSerilizer(job_queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK, headers={"Total-Count": count})


class GetPerfDataPhoneModelData(APIView):

    def get(self, request):
        """
        Get Perf Data page phone_model and rom_version parameter
        Author: Goufuqiang
        Date: 2020/8/27
        :param request:
        :return:
        """
        serializer = CheckoutTboardSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        tboard_id_list = serializer.validated_data.get('tboard', [])
        job_id_list = serializer.validated_data.get('job', [])
        rds_queryset = Rds.objects.filter(tboard_id__in=tboard_id_list, job_id__in=job_id_list)
        phone_model_list = list(rds_queryset.distinct().values_list('phone_model', 'phone_model__phone_model_name'))
        results = []
        for phone_model_id, phone_model_name in phone_model_list:
            if phone_model_id:
                rom_version = rds_queryset.filter(
                    phone_model_id=phone_model_id
                ).distinct(
                    'rom_version_id', 'rom_version__version'
                ).order_by('rom_version__version').values('rom_version_id', 'rom_version__version')

                result = {'phone_model_name': phone_model_name, 'phone_model_id': phone_model_id,
                          'rom_version_data': list(rom_version)}
                results.append(result)
        return Response(results, status=status.HTTP_200_OK)


class GetPerfDataChart(APIView):

    def get(self, request):
        """
        Perf Data page line chart
        Author: Goufuqiang
        Date: 2020/8/28
        :param request:
        :return:
        """
        serializer = CheckoutTboardSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        tboard_id_list = serializer.validated_data.get('tboard', [])
        job_id_list = serializer.validated_data.get('job', [])
        model_rom_version_list = serializer.validated_data.get('model_rom_version', [])
        sizer = serializer.validated_data.get('sizer', None)
        rds_queryset = Rds.objects.filter(tboard_id__in=tboard_id_list, job_id__in=job_id_list,
                                          job_assessment_value='0')
        if sizer == 'box_chart':
            results = {'x_data': [], 'y_data': []}
        else:
            results = []
        for model_rom_version in model_rom_version_list:
            phone_model_id, rom_version_id = tuple(model_rom_version)
            rds_data = rds_queryset.filter(
                phone_model_id=phone_model_id,
                rom_version_id=rom_version_id
            ).values_list('job_duration', 'phone_model__phone_model_name', 'rom_version__version')
            if rds_data:
                job_duration_list = [duration[0] for duration in list(rds_data)]
                _, phone_model_name, rom_version_name = rds_data[0]
                job_duration_list = checkout_np_type_data(job_duration_list)
                median = 0
                avg = 0
                max_data = 0
                if job_duration_list:
                    # median
                    median = round(np.median(job_duration_list), 3)
                    # avg
                    avg = round(np.mean(job_duration_list), 3)
                    # max
                    max_data = max(job_duration_list)
                # pass num
                success_num = rds_queryset.filter(
                    phone_model_id=phone_model_id,
                    rom_version_id=rom_version_id
                ).aggregate(success_num=Count('job_assessment_value'))['success_num']
                # failed num
                failed_num = Rds.objects.filter(
                    tboard_id__in=tboard_id_list, job_id__in=job_id_list, job_assessment_value='1',
                    phone_model_id=phone_model_id, rom_version_id=rom_version_id
                ).aggregate(failed_num=Count('job_assessment_value'))['failed_num']
                if sizer == 'box_chart':
                    # box chart
                    results['x_data'].append(f'{phone_model_name}/{rom_version_name}')
                    results['y_data'].append(job_duration_list)
                elif sizer == 'table_chart':
                    # table chart
                    result = {'phone_model_name': phone_model_name, 'rom_version_name': rom_version_name,
                              'median': median, 'avg': avg, 'max': max_data, 'success_num': success_num,
                              'failed_num': failed_num}
                    results.append(result)
                else:
                    # line chart
                    try:
                        phone_model = PhoneModel.objects.get(id=phone_model_id)
                        rom_version = RomVersion.objects.get(id=rom_version_id)
                    except Exception as e:
                        return Response(e, status=status.HTTP_400_BAD_REQUEST)
                    results.append([f'{phone_model_name}/{rom_version_name}', median, phone_model_id, rom_version_id,
                                    phone_model.phone_model_name, rom_version.version])
        return Response(results, status=status.HTTP_200_OK)


class GetSingleDevicePerfTableData(APIView):

    def get(self, request):
        """
        Single device table data
        Author: Goufuqiang
        Date: 2020/9/15
        :param request:
        :return:
        """
        serializer = CheckoutTboardSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        tboard_id_list = serializer.validated_data.get('tboard', [])
        job_id_list = serializer.validated_data.get('job', [])
        rds_queryset = Rds.objects.filter(tboard_id__in=tboard_id_list, job_id__in=job_id_list)
        failed_num = rds_queryset.filter(job_assessment_value='1').count()
        rds_queryset = rds_queryset.filter(job_assessment_value='0')
        phone_model = serializer.validated_data.get('phone_model_obj', None)
        rom_version = serializer.validated_data.get('rom_version_obj', None)
        if phone_model and rom_version:
            rds_queryset = rds_queryset.filter(phone_model=phone_model, rom_version=rom_version)
        job_duration_list = list(rds_queryset.values_list('job_duration', flat=True))
        job_duration_list = checkout_np_type_data(job_duration_list)
        median = 0
        avg = 0
        max_data = 0
        if job_duration_list:
            # median
            median = round(np.median(job_duration_list), 2)
            # avg
            avg = round(np.mean(job_duration_list), 2)
            # max
            max_data = max(job_duration_list)
        results = {'avg': avg, 'max': max_data, 'median': median, 'failed_num': failed_num,
                   'success_num': rds_queryset.count()}
        return Response(results, status=status.HTTP_200_OK)


class GetPerfDataBarChart(APIView):

    def get(self, request):
        """
        ApiDescription: Bar Chart
        Author: Goufuqiang
        Date: 2020/9/15
        """
        serializer = CheckoutTboardSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        tboard_id_list = serializer.validated_data.get('tboard', [])
        job_id_list = serializer.validated_data.get('job', [])
        device_queryset = serializer.validated_data.get('devices', [])
        rds_queryset = Rds.objects.filter(tboard_id__in=tboard_id_list, job_id__in=job_id_list,
                                          job_assessment_value='0', device__in=device_queryset)
        phone_model = serializer.validated_data.get('phone_model_obj', None)
        rom_version = serializer.validated_data.get('rom_version_obj', None)
        if phone_model and rom_version:
            rds_queryset = rds_queryset.filter(phone_model=phone_model, rom_version=rom_version)
        job_duration_list = list(rds_queryset.values_list('job_duration', flat=True))
        job_duration_list = [job_duration for job_duration in job_duration_list if job_duration is not None]
        results = []
        if job_duration_list:
            results = statistics_number_range(job_duration_list)
            results = [[data, results[data]] for data in results]
        return Response(results, status=status.HTTP_200_OK)


class GetPerfDataTimeBarChart(APIView):
    def get(self, request):
        serializer = PerfViewSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        device_queryset = serializer.validated_data.get('devices', [])
        rdss = Rds.objects.filter(tboard_id=serializer.validated_data.get("tboard"),
                                  job_id=serializer.validated_data.get("job"),
                                  device__in=device_queryset,
                                  job_duration__gt=0,
                                  job_duration__isnull=False).all()
        data = [[timezone.datetime.strftime(timezone.localtime(rds.start_time), "%Y-%m-%d %H:%M:%S"),
                 rds.job_duration] for rds in rdss]
        return Response(data, status=status.HTTP_200_OK)


class GetXlsData(APIView):
    def get(self, request):
        serializer = XLSTboradSerilizer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        tboard = serializer.validated_data.get("tboard")
        job_res_rule = serializer.validated_data.get("job_res_rule", {})

        excel_file_name = f'{tboard.board_name}_{tboard.id}.xlsx'
        file_name = os.path.join(MEDIA_ROOT, RESOURCE_EXCEL_FILE_EXPORT, excel_file_name)
        rds_queryset = Rds.objects.filter(tboard=tboard)
        job_duration_rds_queryset = rds_queryset.exclude(job_duration__isnull=True)
        writer = pd.ExcelWriter(file_name, engine='openpyxl')
        # att source add
        jobs = tboard.job.all().distinct('job_name')

        self.create_avg_data_sheet(tboard, jobs, writer, job_res_rule)

        self.create_job_sheet(tboard, jobs, writer)

        # 有 job_duration 数据才写入excel文件
        if job_duration_rds_queryset.exists():
            self.create_job_duration_sheet(job_duration_rds_queryset, writer)
        writer.save()
        writer.close()
        # response = FileResponse(open(file_name, "rb"))
        # response['Content-Type'] = 'application/octet-stream'
        # response['Content-Disposition'] = f'attachment;filename="{escape_uri_path(tboard.board_name)}.xlsx"'
        export_file_name = os.path.join(RESOURCE_EXCEL_FILE_EXPORT, excel_file_name)
        return ReefResponse(data=export_file_name)


    def create_job_duration_sheet(self, job_duration_rds_queryset, writer):
        perf_rds = job_duration_rds_queryset.annotate(
            time=Cast(
                TruncSecond('start_time', tzinfo=timezone.get_default_timezone()), CharField(max_length=19)
            )
        ).values_list(
            "job__job_name", "device__device_name", "id", "time", "job_assessment_value", "job_duration"
        ).order_by("job__job_name", "time")
        df = pd.DataFrame(perf_rds, columns=("用例名称", "设备名称", "RDSID", "开始时间", "RDS结果", "测试结果/s"))
        # 此处engine如果使用xlsxwriter后面的样式设置需要换对应的写法。

        df.to_excel(writer, sheet_name='启动时间', index=False)

    def set_xls_format(self, sheet_name, tBoard, worksheet):
        # 先插入空行
        for i in range(1, 4):
            worksheet.insert_rows(i)
        font = Font(name='Times New Roman', bold=True)
        # 插入表头信息
        worksheet.cell(1, 1, value=f'测试任务:{tBoard.board_name}').font = font
        worksheet.cell(2, 1, value=f'测试设备:{sheet_name}').font = font
        worksheet.cell(3, 1, value=f'导出时间:{timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")}').font = font
        for cell in worksheet["C"]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # 设置单元格颜色
        for i in range(1, 4):
            worksheet.cell(4, i).fill = PatternFill(fill_type='solid', fgColor="97a675")
            worksheet.cell(i, 1).fill = PatternFill(fill_type='solid', fgColor="d6ead4")
            worksheet.cell(i, 2).fill = PatternFill(fill_type='solid', fgColor="d6ead4")
            worksheet.cell(i, 3).fill = PatternFill(fill_type='solid', fgColor="d6ead4")
        # 调整默认列宽
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 21
        worksheet.column_dimensions['C'].width = 12

    def create_avg_data_sheet(self, tboard, jobs, writer, job_res_rule):
        device_queryset = tboard.device.all().distinct('device_label')
        for device in device_queryset:
            df_data = []
            for job in jobs:
                # 这里不能按job分组统计，要获取rds_dict 内容来判断是否有recognize_words数据
                rds_queryset = Rds.objects.filter(device=device, tboard=tboard, job=job)
                recognize_list = []
                # 启动时间类型统计
                line_data = {'用例名称': job.job_name, '数据名称': '启动时间'}
                job_duration_avg = rds_queryset.exclude(job_duration__isnull=True).aggregate(
                    job_duration_avg=Avg('job_duration')
                ).get('job_duration_avg', 0)
                # 不是时间测试类型job跳过
                if job_duration_avg is not None:
                    job_duration_avg = Decimal(f'{job_duration_avg}').quantize(Decimal('0.000'), rounding=ROUND_HALF_UP)
                    rule = job_res_rule.get(job.job_name, {}).get('启动时间', None)
                    # 没有给出规则不进行判断
                    if rule == '' or not rule:
                        line_data['平均值'] = job_duration_avg
                        df_data.append(line_data)
                        continue
                    if rule:
                        rule, show_rule = self.handle_rules(rule)
                        try:
                            code = f'{job_duration_avg}{rule}'
                            res_item = self.eval_code(code)
                            line_data.update({'平均值': job_duration_avg, '标准': show_rule, '结果': res_item})
                        except Exception as e:
                            line_data.update({'平均值': job_duration_avg, '标准': show_rule})
                        df_data.append(line_data)
                        continue

                for rds in rds_queryset:
                    if rds.job_duration is None:
                        rds_dict = rds.rds_dict
                        for dict_item in rds_dict:
                            if not isinstance(dict_item, dict):
                                continue
                            recognize_words = dict_item.get('recognize_words', None)
                            if recognize_words:
                                for recognize_word in recognize_words:
                                    recognize_list.append(recognize_word)
                # 统计所有rds recognize_words 中的每一项数据结果值到表中
                recognize_words_df = pd.DataFrame(recognize_list)
                # 空字符使用NaN 填充
                recognize_words_df = recognize_words_df.replace('', np.NaN)
                for column in recognize_words_df.columns:
                    try:
                        recognize_word_avg = recognize_words_df[column].astype('float').mean()
                        if pd.isnull(recognize_word_avg):
                            recognize_word_avg = ''
                        else:
                            recognize_word_avg = Decimal(f'{recognize_word_avg}').quantize(Decimal('0.000'), rounding=ROUND_HALF_UP)
                    except Exception as e:
                        recognize_word_avg = ''
                    rule = job_res_rule.get(job.job_name, {}).get(column, None)
                    if rule is None or rule == '' or recognize_word_avg == '':
                        line_data = {'用例名称': job.job_name, '数据名称': f'{column}', '平均值': recognize_word_avg}
                    else:
                        rule, show_rule = self.handle_rules(rule)
                        try:
                            code = f'{recognize_word_avg}{rule}'
                            res_item = self.eval_code(code)
                            line_data = {'用例名称': job.job_name, '数据名称': f'{column}',
                                         '平均值': recognize_word_avg, '标准': show_rule, '结果': res_item}
                        except Exception as e:
                            line_data = {'用例名称': job.job_name, '数据名称': f'{column}',
                                         '平均值': recognize_word_avg, '标准': show_rule}
                    df_data.append(line_data)
            # 每个设备一个sheet
            sheet_name = f'{device.device_name}统计'
            self.create_df_data_to_excel(
                df_data, writer, sys._getframe().f_code.co_name, sheet_name, tboard=tboard
            )

            self.handel_cell_equal_sign_data(df_data, writer, sheet_name)

    def handel_cell_equal_sign_data(self, df_data, writer, sheet_name):
        # 单元格内以等于号好开头会被识别为函数，这里讲数据处理为str
        ws = writer.sheets[f'{sheet_name}']
        column_word = None
        att_source_pd = pd.DataFrame(df_data)
        for index, column in enumerate(att_source_pd.columns):
            if column == '标准':
                column_word = openpyxl.utils.get_column_letter(index + 1)
        if column_word:
            for cell in ws[column_word]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.data_type == 'f':
                    cell.data_type = 's'

    def create_job_sheet(self, tboard, jobs, writer):
        rds_queryset = Rds.objects.filter(tboard=tboard)
        for job in jobs:
            df_data = []
            # source_dict = collections.defaultdict(list)
            for rds in rds_queryset.filter(job=job):
                rds_dict = rds.rds_dict
                recognize_words_dict = {}
                for dict_item in rds_dict:
                    if not isinstance(dict_item, dict):
                        continue
                    recognize_words = dict_item.get('recognize_words', None)
                    if recognize_words:
                        for recognize_word in recognize_words:
                            recognize_words_dict.update(recognize_word)
                # rds dict 没有 recognize_words 数据，跳过
                if not recognize_words_dict:
                    continue
                start_time = date_format_transverter(rds.start_time)
                device_name = '' if getattr(rds, 'device', None) is None else getattr(rds, 'device').device_name
                line_data = {'设备名称': device_name, 'RDSID': rds.id, '开始时间': start_time, 'RDS结果': rds.job_assessment_value}
                line_data.update(recognize_words_dict)
                df_data.append(line_data)
            if df_data:
                self.create_df_data_to_excel(df_data, writer, sys._getframe().f_code.co_name, job.job_name, job, tboard, )

    def create_df_data_to_excel(self, df_data, writer, func_name, sheet_name, job=None, tboard=None):
        try:
            att_source_pd = pd.DataFrame(df_data)
            # att_source_pd.fillna('', inplace=True)
            att_source_pd.to_excel(writer, index=False, sheet_name=f'{sheet_name}')
            # raise ValueError('This is value error!!!')
        except Exception as e:
            writer.close()
            return reef_400_response(message=f"init DataFrame or write excel file error: {e}  \n"
                                             f"func name: {func_name} \n"
                                             f"tboard: {tboard.id}  \n"
                                             f"job: {job.id}  \n"
                                             f"att_source: {df_data}\n")

    def eval_code(self, code):
        try:
            if eval(code):
                res_item = 'pass'
            else:
                res_item = 'fail'
            return res_item
        except Exception as e:
            raise SyntaxError

    def handle_rules(self, rule):
        # =5 --> ==5
        if '=' in rule and rule.startswith('='):
            return rule.replace('=', '=='), rule
        return rule, rule


class PerfDataPreviewView(AutoExecuteSerializerGenericAPIView):

    serializer_class = PerfDataPreviewSerializer

    def get(self, request):
        serializer = self.execute(request, action='get')
        tboard = serializer.validated_data.get('tboard')
        jobs = tboard.job.all().distinct('job_name')
        rep_res = []
        for job in jobs:
            rds_queryset = Rds.objects.filter(job=job, tboard=tboard)
            for rds in rds_queryset:
                if rds.job_duration is not None:
                    item = {'job_name': job.job_name, 'data_name': '启动时间'}
                    if item not in rep_res:
                        rep_res.append(item)
                else:
                    rds_dict = rds.rds_dict
                    for dict_item in rds_dict:
                        if not isinstance(dict_item, dict):
                            continue
                        recognize_words = dict_item.get('recognize_words', None)
                        if recognize_words:
                            for recognize_word in recognize_words:
                                item = {'job_name': job.job_name, 'data_name': list(recognize_word.keys())[0]}
                                if item not in rep_res:
                                    rep_res.append(item)
        return ReefResponse(data=rep_res)


def statistics_number_range(sample_data_list: list):
    # 每隔0.5为一个范围
    # sample_data_list.sort()
    # result_list = []
    # for data in sample_data_list:
    #     int_data = data * 100
    #     first_num = int_data // 10
    #     last_num = data % 10
    #     if last_num <= 5 and last_num >=0:
    #         max_range = (first_num + 1) * 10 - 5
    #         min_range = first_num * 10
    #     else:
    #         min_range = first_num * 10 + 6
    #         max_range = (first_num + 1) * 10
    #     result_list.append(f'{round(min_range / 100, 2)}-{round(max_range / 100, 2)}')
    # return dict(Counter(result_list))

    # 校验数据保留3位小数
    sample_data_list = [round(i, 4) for i in sample_data_list if i > 0]
    results = {}
    if sample_data_list:
        if len(sample_data_list) == 1:
            data = sample_data_list[0]
            min_num = math.floor(1000 * data) / 1000
            math_num = math.ceil(1000 * data) / 1000
            results[f'{min_num}-{math_num}'] = 1
        else:
            # qcut precision 默认精度为3
            categories = pd.cut(sample_data_list, 5, precision=4, duplicates='drop')
            for interval, count in categories.value_counts().items():
                # 左边界 右边界  数量
                results[f'{round(interval.left, 4)}-{round(interval.right, 4)}'] = count
    return results


def checkout_np_type_data(data_list: list):
    if not all(data_list):
        data_list = list(filter(filter_data, data_list))
    return data_list


def filter_data(data):
    # 删除 None,str类型数据
    if data is not None and not isinstance(data, str):
        return data
